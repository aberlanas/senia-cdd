#!/usr/bin/python3

import sys
import os
import random
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import openpyxl
import argparse

from openpyxl import load_workbook
wb= load_workbook(filename = '/home/tic/Escritorio/ocupacionModificadoTotal.xlsx')
wbflip = Workbook()
ws = wb.active
wsflip = wbflip.active

#celda = ws.cell(column=columna,row=fila,value=")

for fila in range(1,100):
    for columna in range(1,100):
        print("Rellenado la celda : "+str(fila)+","+str(columna)+"")
        
        celda_orig = ws.cell(column=columna,row=fila)
        celda_flip = wsflip.cell(column=fila,row=columna)
        celda_flip.value=celda_orig.value
        
        if (celda_orig.fill.start_color.index == '00000000'):
            aux_color = openpyxl.styles.colors.Color(rgb='FFFFFFFF')
        else:
            aux_color = openpyxl.styles.colors.Color(celda_orig.fill.start_color.index)
       
        aux_pattern = PatternFill(fill_type='solid', fgColor=aux_color)
       
        print(aux_pattern)
        
        try:
            celda_flip.fill=aux_pattern
        except Exception as e:
            print(e)


wbflip.save("patata.xlsx")

sys.exit(0)
