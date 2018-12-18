#!/usr/bin/env python3
#
#

# Dependencias : python3-openpyxl


import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import load_workbook
import openpyxl

print(" * Welcome to python horarios")

class Sesion:

    def __init__(self,dia_semana,sesion_orden,aula,grupo,materia):
        self.dia_semana = dia_semana
        self.sesion_orden = sesion_orden
        self.aula = aula
        self.grupo = grupo
        self.materia = materia

    def print_sesion(self):
        print(" * "+self.dia_semana+" : "+self.sesion_orden+" - "+self.grupo+" - "+self.aula)


class Profesor:

    def __init__(self,nombre,apellido):
        self.nombre=nombre
        self.apellido = apellido
        self.listaSesiones = []

# Cargamos el fichero de origen
horarios_file="ocupacion.xlsx"
horarios_profes="patatasfritasconmostaza.xlsx"

wb_orig = load_workbook(filename = horarios_file)
ws_orig = wb_orig.active

wb_dest = Workbook()

profesores = []


## Metodos para trabajar con los profesores

def busca_y_anyade(nombre,apellido):
    aux_profe = Profesor(nombre,apellido)
    found = False
    for profe in profesores:
        if profe.nombre == aux_profe.nombre and profe.apellido == aux_profe.apellido:
            found = True
            return profe
    if not found:
        profesores.append(aux_profe)
        return aux_profe







## 
for columna in range(2,100):
    for fila in range(2,100):
        celda_aux = ws_orig.cell(column=columna,row=fila)
        if (celda_aux.value == None):
            pass
        else:
            print("--------------")
            #print(celda_aux.value)

            # Trabajamos esta celda
            hora = ws_orig.cell(column=1,row=fila).value.strip('\r\n')

            # Dirty hack
            dia_hora = hora[:1]
            hora_hora = hora.split(":",1)[1]
            
            datos_celda = celda_aux.value.replace("\n"," ").split(" ")
            print(datos_celda)

            ## Dirty hack?
            p_nombre_aux = datos_celda[2]

            p_apellidos_aux =""
            longitud = len(datos_celda)
            long_aux = 3
            while(long_aux < longitud):
                
                p_apellidos_aux+=datos_celda[long_aux]
                long_aux=long_aux+1
            
            
            
            p_grupo = datos_celda[0]
            p_materia = datos_celda[1]
            
            aula = ws_orig.cell(column=columna,row=1).value.split("[")[0].rstrip()

            print(dia_hora+":"+hora_hora+" -> "+aula)

            profe_aux = busca_y_anyade(p_nombre_aux,p_apellidos_aux)

            sesion_aux = Sesion(dia_hora,hora_hora,aula,p_grupo,p_materia)
            profe_aux.listaSesiones.append(sesion_aux)

# Estilos
thin = Side(border_style="thin", color="000000")
border = Border(top=thin, left=thin, right=thin, bottom=thin)
al = Alignment(horizontal="center", vertical="center")

# Ahora rellenamos el excel
for profe in profesores:
    print("++++++++++++++++++")
    print(" * "+profe.nombre+" . "+ profe.apellido+":"+str(len(profe.listaSesiones)))

    # Create sheet
    ws_dest_profe = wb_dest.create_sheet(profe.nombre+"."+profe.apellido)

    # Profesor
    celda=ws_dest_profe.cell(column=1,row=1)
    celda.value=profe.nombre+"."+profe.apellido
    celda.style='Headline 3'

    ws_dest_profe.column_dimensions["A"].width=15

    for col in ["B","C","D","E","F"]:
        ws_dest_profe.column_dimensions[col].width=15
        ws_dest_profe.column_dimensions[col].alignment = Alignment(horizontal='center')
        ws_dest_profe.column_dimensions[col].alignment = Alignment(vertical='center')

    for row in range(1,20):
        ws_dest_profe.row_dimensions[row].height=40
        
    
    # Horas
    columna=1
    fila=2
    horas_docencia =["8:00-8:55","8:55-9:50","9:50-10:45","10:45-11:05","11:05-12:00","12:00-12:55","12:55-13:50","13:50-14:10","14:10-15:05","15:05-16:00","16:00-16:55","16:55-17:15","17:15-18:10","18:10-19:05","19:05-20:00","20:00-20:55"]
    for hora in horas_docencia:
        celda = ws_dest_profe.cell(column=columna,row=fila)
        celda.value=hora
        celda.style='Headline 2'
        fila=fila+1

    # Dias
    fila=2
    columna=2
    for dia in ["L","M","X","J","V"]:
        celda = ws_dest_profe.cell(column=columna,row=1)
        celda.value=dia

        celda.style='Headline 1'
        celda.alignment=Alignment(vertical="center")
        celda.alignment=Alignment(horizontal="center")

    
        for sesion in profe.listaSesiones:
            print(sesion.dia_semana+","+sesion.sesion_orden+","+sesion.aula+","+sesion.grupo+","+sesion.materia)
            if (sesion.dia_semana==dia):
                celda_sesion=ws_dest_profe.cell(column=columna,row=int(sesion.sesion_orden)+1)
                celda_sesion.value="["+sesion.aula+"]\n"+sesion.grupo+":\n"+sesion.materia

        for celda_tabla_n in range(2,18):
            celda_tabla = ws_dest_profe.cell(column=columna,row=celda_tabla_n)
            celda_tabla.border = border
            celda_tabla.alignment = al

        
        # Incrementamos la columna
        columna=columna+1

    ws_dest_profe.page_margins.top=1
    ws_dest_profe.page_margins.bottom=1
    ws_dest_profe.page_margins.left=0.5
    ws_dest_profe.page_margins.right=0.5
    
wb_dest.save("patatasfritas.xlsx")




