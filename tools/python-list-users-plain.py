#!/usr/bin/python3
 # -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl import load_workbook

listFile = "/home/aberlanas/temp/GRUPOS_2021.xlsx"
fgroupsPath = "/tmp/listGroups.txt"

wb = load_workbook(listFile)
fgroups = open(fgroupsPath,'x')

for sheet_name in wb.sheetnames:

   sheet = wb[sheet_name]

   print( "" )
   print( "# Grup : " + sheet['A1'].value)
   fgroups.write(sheet['A1'].value+"\n")
   print( "")

   print("|Nº| Nom |Cognoms|")
   print("|:-:|:---|:------|")
   
   numA = 1
   for row in sheet.iter_rows(min_row=4,max_row=24):
       nombre = None

       try:
           nombre = row[0].value
       except Exception as e:
           print(str(e))
           pass

       if nombre != None:
           #print (str(row[0].value))
           nombrel = nombre.split(',')

           try:
              nom = nombrel[1]
              ape = nombrel[0]
           except Exception as e:
              print (str(e))

           nomFinal = ""
           noml = nom.strip().split(' ')
           for auxn in noml:
               nomFinal += auxn.capitalize() + " "

           apelFinal = ""    
           apel = ape.strip().split(' ')

           for auxp in apel:
               apelFinal+=auxp.capitalize()+" "

           print ("| " + str(numA) + " | "+ nomFinal + "|" + apelFinal +"|")
           numA = numA + 1

   print(" ")
   print("---")
   print(" ")
   print("\\newpage ")
   print(" ")

fgroups.close()
